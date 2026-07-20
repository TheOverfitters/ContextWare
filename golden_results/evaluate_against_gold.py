#!/usr/bin/env python3
"""Evaluate the three cloud ACF classifiers against the human golden standard.

This script compares the predictions of three models

    * glm-5.2:cloud
    * kimi-k2.7-code:cloud
    * qwen3.5:cloud

against the human-annotated golden standard (``labels_gold`` and
``maintenance_gold``) and reports **Precision, Recall and F1** for each model on
two axes:

    1. Category axis (multi-label)   -> ``labels_gold`` vs ``category_scores``
    2. Maintenance axis (single-label) -> ``maintenance_gold`` vs
       ``maintenance.primary_maintenance_type``

It additionally quantifies the **semantic ambiguity** of the instructions in the
Agent Context Files by measuring the *divergence between the models*: where the
three models disagree the most, the underlying instruction is treated as more
semantically ambiguous. The script reports per-item divergence and shows how
divergence correlates with the models' error rate against the gold standard.

The whole thing is pure standard library (plots are optional and only drawn when
matplotlib is importable), so it can run inside the existing acf-analysis
environment without extra dependencies.

Usage (from the repository root)::

    python golden_results/evaluate_against_gold.py

Custom paths / threshold::

    python golden_results/evaluate_against_gold.py \
        --gold golden_results/data/gold_sample.csv \
        --predictions acf-analysis/acf-outputs \
        --threshold 0.5 \
        --out golden_results/results
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import math
import os
from collections import Counter
from itertools import combinations
from pathlib import Path

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent

# Directory name (on disk) -> human readable model id used in the reports.
# NOTE: gemma4:31b-cloud was the original "reference" model used to seed the
# golden standard, so its agreement with the gold enjoys a head start; read its
# scores as a peer-with-context rather than a fully independent comparison.
MODELS: dict[str, str] = {
    "glm-5.2_cloud": "glm-5.2:cloud",
    "kimi-k2.7-code_cloud": "kimi-k2.7-code:cloud",
    "qwen3.5_cloud": "qwen3.5:cloud",
    "gemma4_31b-cloud": "gemma4:31b-cloud",
}

# Maintenance-gold values are hand typed, so normalise obvious noise:
#   - a trailing "?" marks annotator uncertainty
#   - "prefective" is a misspelling of "perfective"
MAINTENANCE_FIXUPS = {"prefective": "perfective"}

# The six leaf maintenance types the classifier can emit.
MAINTENANCE_TYPES = [
    "corrective",
    "preventive",
    "adaptive (correction)",
    "adaptive (enhancement)",
    "additive",
    "perfective",
]


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #
def load_category_names(path: Path) -> list[str]:
    """Ordered canonical category names from categories.json."""
    data = json.loads(path.read_text(encoding="utf-8"))
    items = data.get("categories", data) if isinstance(data, dict) else data
    names: list[str] = []
    for it in items or []:
        name = it.get("name") if isinstance(it, dict) else it
        if isinstance(name, str) and name:
            names.append(name)
    return names


def normalise_maintenance(value: str) -> str:
    """Canonicalise a maintenance label (strip uncertainty '?', fix typos)."""
    v = (value or "").strip().rstrip("?").strip().lower()
    v = MAINTENANCE_FIXUPS.get(v, v)
    return v


def _iter_gold_rows(path: Path):
    """Yield ``{diff_id, labels_gold, maintenance_gold}`` dicts from the gold
    file, transparently handling both ``.xlsx`` (the reference format) and the
    semicolon-delimited, BOM-prefixed ``.csv`` export."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl  # local import: only needed for the Excel path
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        col = {name: header.index(name) for name in header}
        for raw in it:
            def cell(name):
                i = col.get(name)
                return "" if i is None or raw[i] is None else str(raw[i])
            yield {
                "diff_id": cell("diff_id"),
                "labels_gold": cell("labels_gold"),
                "maintenance_gold": cell("maintenance_gold"),
            }
    else:
        with path.open(encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                yield row


def load_gold(path: Path, categories: list[str]) -> dict[str, dict]:
    """Return ``{diff_id: {"labels": frozenset, "maintenance": str}}``.

    Reads the ``.xlsx`` reference (or a ``.csv`` export). Category tokens are
    validated against the canonical vocabulary; unknown tokens are reported so
    annotation typos surface instead of being silently dropped.
    """
    catset = set(categories)
    unknown: Counter[str] = Counter()
    gold: dict[str, dict] = {}
    for row in _iter_gold_rows(path):
        diff_id = (row.get("diff_id") or "").strip()
        if not diff_id:
            continue
        labels = set()
        for tok in (row.get("labels_gold") or "").split(";"):
            tok = tok.strip()
            if not tok:
                continue
            if tok in catset:
                labels.add(tok)
            else:
                unknown[tok] += 1
        gold[diff_id] = {
            "labels": frozenset(labels),
            "maintenance": normalise_maintenance(row.get("maintenance_gold", "")),
        }
    if unknown:
        print("  [warn] gold label tokens not in canonical vocabulary:",
              dict(unknown))
    return gold


def load_model_predictions(pred_dir: Path, model_dir: str) -> dict[str, dict]:
    """Merge every ``primary_*.jsonl`` for a model (later runs win)."""
    files = sorted(glob.glob(str(pred_dir / model_dir / "primary_*.jsonl")))
    by_id: dict[str, dict] = {}
    for fn in files:
        with open(fn, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                diff_id = str(rec.get("diff_id", ""))
                if diff_id:
                    by_id[diff_id] = rec
    return by_id


def predicted_category_set(rec: dict, categories: list[str], threshold: float) -> frozenset[str]:
    """Multi-label prediction: categories with score >= threshold."""
    scores = rec.get("category_scores") or {}
    return frozenset(c for c in categories if float(scores.get(c, 0.0)) >= threshold)


def predicted_maintenance(rec: dict) -> str:
    m = rec.get("maintenance") or {}
    return normalise_maintenance(m.get("primary_maintenance_type", ""))


# --------------------------------------------------------------------------- #
# Metric primitives
# --------------------------------------------------------------------------- #
def prf(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    """Precision, Recall, F1 from raw counts (0 when undefined)."""
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f = 2 * p * r / (p + r) if (p + r) else 0.0
    return p, r, f


def jaccard(a: frozenset[str], b: frozenset[str]) -> float:
    if not a and not b:
        return 1.0
    return len(a & b) / len(a | b)


def shannon_entropy_bits(votes: list[str]) -> float:
    n = len(votes)
    if n <= 1:
        return 0.0
    counts = Counter(votes)
    return -sum((c / n) * math.log2(c / n) for c in counts.values())


def pearson(xs: list[float], ys: list[float]) -> float:
    n = len(xs)
    if n < 2:
        return float("nan")
    mx, my = sum(xs) / n, sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    dx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    dy = math.sqrt(sum((y - my) ** 2 for y in ys))
    return num / (dx * dy) if dx and dy else float("nan")


# --------------------------------------------------------------------------- #
# Category axis (multi-label P/R/F1)
# --------------------------------------------------------------------------- #
def evaluate_categories(
    diff_ids: list[str],
    gold: dict[str, dict],
    pred_sets: dict[str, frozenset[str]],
    categories: list[str],
) -> tuple[dict, list[dict]]:
    """Return (overall_metrics, per_category_rows) for one model."""
    # Micro accumulators + per-category accumulators.
    micro_tp = micro_fp = micro_fn = 0
    cat_counts = {c: {"tp": 0, "fp": 0, "fn": 0} for c in categories}

    exact_match = 0
    jaccard_vals: list[float] = []
    hamming_bits = 0

    for d in diff_ids:
        g = gold[d]["labels"]
        p = pred_sets[d]
        tp = len(g & p)
        fp = len(p - g)
        fn = len(g - p)
        micro_tp += tp
        micro_fp += fp
        micro_fn += fn
        exact_match += int(g == p)
        jaccard_vals.append(jaccard(g, p))
        hamming_bits += fp + fn
        for c in categories:
            in_g, in_p = c in g, c in p
            if in_g and in_p:
                cat_counts[c]["tp"] += 1
            elif in_p and not in_g:
                cat_counts[c]["fp"] += 1
            elif in_g and not in_p:
                cat_counts[c]["fn"] += 1

    per_category: list[dict] = []
    macro_p = macro_r = macro_f = 0.0
    wsum = 0
    wp = wr = wf = 0.0
    for c in categories:
        cc = cat_counts[c]
        p, r, f = prf(cc["tp"], cc["fp"], cc["fn"])
        support = cc["tp"] + cc["fn"]  # gold occurrences
        per_category.append({
            "category": c, "precision": p, "recall": r, "f1": f,
            "support": support, "tp": cc["tp"], "fp": cc["fp"], "fn": cc["fn"],
        })
        macro_p += p
        macro_r += r
        macro_f += f
        wp += p * support
        wr += r * support
        wf += f * support
        wsum += support

    n_cat = len(categories)
    micro_p, micro_r, micro_f = prf(micro_tp, micro_fp, micro_fn)
    overall = {
        "n_items": len(diff_ids),
        "micro_precision": micro_p,
        "micro_recall": micro_r,
        "micro_f1": micro_f,
        "macro_precision": macro_p / n_cat,
        "macro_recall": macro_r / n_cat,
        "macro_f1": macro_f / n_cat,
        "weighted_precision": (wp / wsum) if wsum else 0.0,
        "weighted_recall": (wr / wsum) if wsum else 0.0,
        "weighted_f1": (wf / wsum) if wsum else 0.0,
        "exact_match_ratio": exact_match / len(diff_ids),
        "mean_jaccard": sum(jaccard_vals) / len(jaccard_vals),
        "hamming_loss": hamming_bits / (len(diff_ids) * n_cat),
        "hamming_accuracy": 1 - hamming_bits / (len(diff_ids) * n_cat),
    }
    return overall, per_category


# --------------------------------------------------------------------------- #
# Maintenance axis (single-label multiclass P/R/F1)
# --------------------------------------------------------------------------- #
def evaluate_single_label(
    diff_ids: list[str],
    gold_of: dict[str, str],
    pred_of: dict[str, str],
    classes: list[str],
) -> tuple[dict, list[dict]]:
    """Macro/weighted P/R/F1 + accuracy for a single-label axis."""
    # Only score items that carry a gold value on this axis.
    scored = [d for d in diff_ids if gold_of.get(d)]
    counts = {c: {"tp": 0, "fp": 0, "fn": 0} for c in classes}
    correct = 0
    for d in scored:
        g = gold_of[d]
        p = pred_of.get(d, "")
        if g == p:
            correct += 1
            if g in counts:
                counts[g]["tp"] += 1
        else:
            if p in counts:
                counts[p]["fp"] += 1
            if g in counts:
                counts[g]["fn"] += 1

    per_class: list[dict] = []
    macro_p = macro_r = macro_f = 0.0
    wp = wr = wf = 0.0
    wsum = 0
    for c in classes:
        cc = counts[c]
        p, r, f = prf(cc["tp"], cc["fp"], cc["fn"])
        support = cc["tp"] + cc["fn"]
        per_class.append({
            "class": c, "precision": p, "recall": r, "f1": f,
            "support": support, "tp": cc["tp"], "fp": cc["fp"], "fn": cc["fn"],
        })
        macro_p += p
        macro_r += r
        macro_f += f
        wp += p * support
        wr += r * support
        wf += f * support
        wsum += support

    n_cls = len(classes)
    overall = {
        "n_items": len(scored),
        "accuracy": correct / len(scored) if scored else 0.0,
        "macro_precision": macro_p / n_cls,
        "macro_recall": macro_r / n_cls,
        "macro_f1": macro_f / n_cls,
        "weighted_precision": (wp / wsum) if wsum else 0.0,
        "weighted_recall": (wr / wsum) if wsum else 0.0,
        "weighted_f1": (wf / wsum) if wsum else 0.0,
    }
    return overall, per_class


# --------------------------------------------------------------------------- #
# Inter-model divergence  ==  semantic ambiguity of the instruction
# --------------------------------------------------------------------------- #
def evaluate_divergence(
    diff_ids: list[str],
    gold: dict[str, dict],
    cat_sets: dict[str, dict[str, frozenset[str]]],
    maint_pred: dict[str, dict[str, str]],
) -> tuple[list[dict], dict]:
    """Per-item divergence between the three models and its link to gold error.

    For every instruction we compute:
      * ``category_divergence`` - mean pairwise Jaccard *distance* between the
        three predicted category sets (0 = identical, 1 = disjoint).
      * ``primary_category_entropy`` - Shannon entropy (bits) of the three
        primary-category votes.
      * ``maintenance_entropy`` - Shannon entropy (bits) of the three
        maintenance votes.
      * ``mean_model_error`` - mean over the 3 models of the multi-label
        Jaccard *distance* to the gold category set (how wrong the models are).

    The Pearson correlation between ``category_divergence`` and
    ``mean_model_error`` quantifies the intuition that ambiguous instructions
    (high inter-model divergence) are also the ones the models get wrong.
    """
    models = list(cat_sets.keys())
    rows: list[dict] = []
    for d in diff_ids:
        sets = [cat_sets[m][d] for m in models]
        pair_jd = [1 - jaccard(a, b) for a, b in combinations(sets, 2)]
        cat_div = sum(pair_jd) / len(pair_jd)

        maint_votes = [maint_pred[m][d] for m in models]
        maint_ent = shannon_entropy_bits(maint_votes)
        maint_disagree = 1 - (max(Counter(maint_votes).values()) / len(models))

        g = gold[d]["labels"]
        model_err = [1 - jaccard(cat_sets[m][d], g) for m in models]
        mean_err = sum(model_err) / len(model_err)

        rows.append({
            "diff_id": d,
            "category_divergence": cat_div,
            "maintenance_entropy": maint_ent,
            "maintenance_disagreement": maint_disagree,
            "mean_model_error_vs_gold": mean_err,
            "unanimous_categories": int(cat_div == 0.0),
            "unanimous_maintenance": int(len(set(maint_votes)) == 1),
        })

    cat_div_vals = [r["category_divergence"] for r in rows]
    err_vals = [r["mean_model_error_vs_gold"] for r in rows]
    maint_ent_vals = [r["maintenance_entropy"] for r in rows]

    summary = {
        "n_items": len(rows),
        "mean_category_divergence": sum(cat_div_vals) / len(rows),
        "mean_maintenance_entropy": sum(maint_ent_vals) / len(rows),
        "share_unanimous_categories": sum(r["unanimous_categories"] for r in rows) / len(rows),
        "share_unanimous_maintenance": sum(r["unanimous_maintenance"] for r in rows) / len(rows),
        "corr_divergence_vs_error": pearson(cat_div_vals, err_vals),
    }

    # Stratify gold error by divergence tercile to make the ambiguity/error
    # link readable at a glance.
    ordered = sorted(rows, key=lambda r: r["category_divergence"])
    n = len(ordered)
    thirds = {
        "low_divergence": ordered[: n // 3],
        "mid_divergence": ordered[n // 3: 2 * n // 3],
        "high_divergence": ordered[2 * n // 3:],
    }
    summary["error_by_divergence_tercile"] = {
        k: (sum(r["mean_model_error_vs_gold"] for r in grp) / len(grp)) if grp else 0.0
        for k, grp in thirds.items()
    }
    return rows, summary


# --------------------------------------------------------------------------- #
# Output helpers
# --------------------------------------------------------------------------- #
def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def round_row(row: dict, ndigits: int = 4) -> dict:
    return {k: (round(v, ndigits) if isinstance(v, float) else v) for k, v in row.items()}


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--gold", type=Path,
                    default=REPO_ROOT / "golden-standard" / "gold-sample" / "gold_sample_blind.xlsx",
                    help="Golden-standard reference (.xlsx reference, or a .csv export).")
    ap.add_argument("--predictions", type=Path, default=REPO_ROOT / "acf-analysis" / "acf-outputs",
                    help="Directory containing one sub-folder of primary_*.jsonl per model.")
    ap.add_argument("--categories", type=Path, default=REPO_ROOT / "categories.json",
                    help="Canonical category vocabulary.")
    ap.add_argument("--threshold", type=float, default=0.5,
                    help="Score threshold for the multi-label category prediction.")
    ap.add_argument("--out", type=Path, default=SCRIPT_DIR / "results",
                    help="Output directory for the CSV / JSON reports.")
    args = ap.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)
    categories = load_category_names(args.categories)

    print(f"Loading gold standard: {args.gold}")
    gold = load_gold(args.gold, categories)
    print(f"  {len(gold)} gold-annotated instructions")

    # Load predictions for the three study models.
    preds: dict[str, dict[str, dict]] = {}
    for mdir in MODELS:
        preds[mdir] = load_model_predictions(args.predictions, mdir)
        print(f"  {MODELS[mdir]}: {len(preds[mdir])} predictions")

    # Only evaluate instructions that every model predicted and that carry gold.
    diff_ids = sorted(
        d for d in gold
        if all(d in preds[m] for m in MODELS)
    )
    print(f"Evaluating {len(diff_ids)} instructions covered by gold + all "
          f"{len(MODELS)} models (threshold={args.threshold})\n")

    # Pre-compute predicted category sets and maintenance labels per model.
    cat_sets = {
        m: {d: predicted_category_set(preds[m][d], categories, args.threshold) for d in diff_ids}
        for m in MODELS
    }
    maint_pred = {
        m: {d: predicted_maintenance(preds[m][d]) for d in diff_ids}
        for m in MODELS
    }

    # ---- Category axis ---------------------------------------------------- #
    cat_overall_rows: list[dict] = []
    cat_percat_rows: list[dict] = []
    for m in MODELS:
        overall, per_cat = evaluate_categories(diff_ids, gold, cat_sets[m], categories)
        overall = {"model": MODELS[m], **overall}
        cat_overall_rows.append(round_row(overall))
        for r in per_cat:
            cat_percat_rows.append(round_row({"model": MODELS[m], **r}))

    write_csv(args.out / "category_metrics_overall.csv", cat_overall_rows,
              list(cat_overall_rows[0].keys()))
    write_csv(args.out / "category_metrics_per_category.csv", cat_percat_rows,
              list(cat_percat_rows[0].keys()))

    # ---- Maintenance axis (leaf + collapsed base type) -------------------- #
    def base(t: str) -> str:
        return "adaptive" if t.startswith("adaptive") else t

    maint_overall_rows: list[dict] = []
    maint_perclass_rows: list[dict] = []
    for m in MODELS:
        # Leaf-type evaluation.
        gold_leaf = {d: gold[d]["maintenance"] for d in diff_ids}
        pred_leaf = maint_pred[m]
        overall, per_cls = evaluate_single_label(diff_ids, gold_leaf, pred_leaf, MAINTENANCE_TYPES)
        maint_overall_rows.append(round_row({"model": MODELS[m], "granularity": "leaf_type", **overall}))
        for r in per_cls:
            maint_perclass_rows.append(round_row({"model": MODELS[m], "granularity": "leaf_type", **r}))

        # Base-type evaluation (adaptive-* collapsed to "adaptive").
        base_classes = ["corrective", "preventive", "adaptive", "additive", "perfective"]
        gold_base = {d: base(gold[d]["maintenance"]) for d in diff_ids}
        pred_base = {d: base(pred_leaf[d]) for d in diff_ids}
        overall_b, per_cls_b = evaluate_single_label(diff_ids, gold_base, pred_base, base_classes)
        maint_overall_rows.append(round_row({"model": MODELS[m], "granularity": "base_type", **overall_b}))
        for r in per_cls_b:
            maint_perclass_rows.append(round_row({"model": MODELS[m], "granularity": "base_type", **r}))

    write_csv(args.out / "maintenance_metrics_overall.csv", maint_overall_rows,
              list(maint_overall_rows[0].keys()))
    write_csv(args.out / "maintenance_metrics_per_class.csv", maint_perclass_rows,
              list(maint_perclass_rows[0].keys()))

    # ---- Divergence / ambiguity ------------------------------------------ #
    div_rows, div_summary = evaluate_divergence(diff_ids, gold, cat_sets, maint_pred)
    write_csv(args.out / "model_divergence_per_item.csv",
              [round_row(r) for r in div_rows], list(div_rows[0].keys()))

    # ---- Summary JSON ----------------------------------------------------- #
    summary = {
        "config": {
            "gold": str(args.gold),
            "predictions": str(args.predictions),
            "threshold": args.threshold,
            "n_items_evaluated": len(diff_ids),
            "models": list(MODELS.values()),
        },
        "category_axis": {r["model"]: {k: v for k, v in r.items() if k != "model"}
                          for r in cat_overall_rows},
        "maintenance_axis": {
            r["model"] + " / " + r["granularity"]: {
                k: v for k, v in r.items() if k not in ("model", "granularity")
            }
            for r in maint_overall_rows
        },
        "semantic_ambiguity": round_row(div_summary),
    }
    (args.out / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    # ---- Console recap ---------------------------------------------------- #
    print("Category axis (multi-label)  P / R / F1")
    print(f"  {'model':<22} {'micro-P':>8} {'micro-R':>8} {'micro-F1':>9} "
          f"{'macro-F1':>9} {'exact':>7} {'Jaccard':>8}")
    for r in cat_overall_rows:
        print(f"  {r['model']:<22} {r['micro_precision']:>8.3f} {r['micro_recall']:>8.3f} "
              f"{r['micro_f1']:>9.3f} {r['macro_f1']:>9.3f} "
              f"{r['exact_match_ratio']:>7.3f} {r['mean_jaccard']:>8.3f}")

    print("\nMaintenance axis (single-label)  Precision / Recall / F1")
    for r in maint_overall_rows:
        print(f"  {r['model']:<22} [{r['granularity']:<9}] acc={r['accuracy']:.3f}  "
              f"macro P/R/F1={r['macro_precision']:.3f}/{r['macro_recall']:.3f}/{r['macro_f1']:.3f}  "
              f"weighted P/R/F1={r['weighted_precision']:.3f}/{r['weighted_recall']:.3f}/{r['weighted_f1']:.3f}")

    print("\nSemantic ambiguity (inter-model divergence)")
    print(f"  mean category divergence      : {div_summary['mean_category_divergence']:.3f}")
    print(f"  mean maintenance entropy (bits): {div_summary['mean_maintenance_entropy']:.3f}")
    print(f"  unanimous on categories        : {div_summary['share_unanimous_categories']:.1%}")
    print(f"  unanimous on maintenance       : {div_summary['share_unanimous_maintenance']:.1%}")
    print(f"  corr(divergence, gold error)   : {div_summary['corr_divergence_vs_error']:.3f}")
    tercile = div_summary["error_by_divergence_tercile"]
    print(f"  gold error by divergence tercile: "
          f"low={tercile['low_divergence']:.3f}  "
          f"mid={tercile['mid_divergence']:.3f}  "
          f"high={tercile['high_divergence']:.3f}")

    # ---- Optional plots --------------------------------------------------- #
    try:
        _draw_plots(args.out, cat_overall_rows, maint_overall_rows,
                    maint_perclass_rows, div_rows, div_summary)
        print(f"\nPlots written to {args.out}")
    except Exception as exc:  # matplotlib missing or headless issue - non-fatal
        print(f"\n[info] plots skipped ({exc.__class__.__name__}: {exc})")

    print(f"\nReports written to {args.out}")


def _draw_plots(out: Path, cat_rows, maint_rows, maint_perclass_rows,
                div_rows, div_summary) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    models = [r["model"] for r in cat_rows]

    # Short class labels for readable tick marks.
    short = {
        "adaptive (correction)": "adapt(corr)",
        "adaptive (enhancement)": "adapt(enh)",
    }

    # 1) Category micro P/R/F1 + Hamming accuracy per model. Accuracy here is the
    #    Hamming accuracy (1 - Hamming loss): the mean fraction of the 16 per-label
    #    present/absent decisions the model gets right - i.e. element-wise, NOT the
    #    stricter exact-match of the whole label set.
    fig, ax = plt.subplots(figsize=(9, 4.5))
    x = range(len(models))
    width = 0.2
    ax.bar([i - 1.5 * width for i in x], [r["micro_precision"] for r in cat_rows], width, label="micro Precision")
    ax.bar([i - 0.5 * width for i in x], [r["micro_recall"] for r in cat_rows], width, label="micro Recall")
    ax.bar([i + 0.5 * width for i in x], [r["micro_f1"] for r in cat_rows], width, label="micro F1")
    ax.bar([i + 1.5 * width for i in x], [r["hamming_accuracy"] for r in cat_rows], width,
           label="Accuracy", color="#777777")
    ax.set_xticks(list(x))
    ax.set_xticklabels(models, rotation=15, ha="right")
    ax.set_ylim(0, 1)
    ax.set_title("Category axis (multi-label, micro) vs golden standard")
    ax.legend(ncol=2)
    fig.tight_layout()
    fig.savefig(out / "category_prf.png", dpi=150)
    plt.close(fig)

    # 2) Maintenance OVERALL (maintenance_metrics_overall.csv): macro
    #    Precision/Recall/F1 + accuracy per model, leaf-type granularity.
    leaf = [r for r in maint_rows if r["granularity"] == "leaf_type"]
    leaf_models = [r["model"] for r in leaf]
    fig, ax = plt.subplots(figsize=(9, 4.5))
    x = range(len(leaf))
    width = 0.2
    ax.bar([i - 1.5 * width for i in x], [r["macro_precision"] for r in leaf], width, label="Precision")
    ax.bar([i - 0.5 * width for i in x], [r["macro_recall"] for r in leaf], width, label="Recall")
    ax.bar([i + 0.5 * width for i in x], [r["macro_f1"] for r in leaf], width, label="F1")
    ax.bar([i + 1.5 * width for i in x], [r["accuracy"] for r in leaf], width, label="Accuracy", color="#777777")
    ax.set_xticks(list(x))
    ax.set_xticklabels(leaf_models, rotation=15, ha="right")
    ax.set_ylim(0, 1)
    ax.set_title("Maintenance axis (single-label, macro) vs golden standard")
    ax.legend(ncol=2)
    fig.tight_layout()
    fig.savefig(out / "maintenance_prf.png", dpi=150)
    plt.close(fig)

    # 3) Maintenance PER-CLASS (maintenance_metrics_per_class.csv): one panel per
    #    model, P/R/F1 across the six maintenance types, support in the labels.
    per = [r for r in maint_perclass_rows if r["granularity"] == "leaf_type"]
    classes = MAINTENANCE_TYPES
    fig, axes = plt.subplots(2, 2, figsize=(13, 8), sharey=True)
    width = 0.27
    xs = range(len(classes))
    for ax, model in zip(axes.flat, leaf_models):
        rows_by_class = {r["class"]: r for r in per if r["model"] == model}
        p = [rows_by_class[c]["precision"] for c in classes]
        r_ = [rows_by_class[c]["recall"] for c in classes]
        f = [rows_by_class[c]["f1"] for c in classes]
        sup = [rows_by_class[c]["support"] for c in classes]
        ax.bar([i - width for i in xs], p, width, label="Precision")
        ax.bar(list(xs), r_, width, label="Recall")
        ax.bar([i + width for i in xs], f, width, label="F1")
        ax.set_title(model)
        ax.set_ylim(0, 1)
        ax.set_xticks(list(xs))
        ax.set_xticklabels([f"{short.get(c, c)}\n(n={s})" for c, s in zip(classes, sup)],
                           rotation=30, ha="right", fontsize=8)
    axes.flat[0].legend(loc="upper right", fontsize=8)
    fig.suptitle("Maintenance axis — per-class Precision / Recall / F1 vs golden standard")
    fig.tight_layout(rect=(0, 0, 1, 0.97))
    fig.savefig(out / "maintenance_per_class.png", dpi=150)
    plt.close(fig)

    # 3) Divergence vs gold error scatter.
    fig, ax = plt.subplots(figsize=(6.5, 5))
    ax.scatter([r["category_divergence"] for r in div_rows],
               [r["mean_model_error_vs_gold"] for r in div_rows], alpha=0.6, s=25)
    ax.set_xlabel("inter-model category divergence (ambiguity)")
    ax.set_ylabel("mean model error vs gold (Jaccard distance)")
    ax.set_title(f"Ambiguity vs error  (Pearson r = {div_summary['corr_divergence_vs_error']:.2f})")
    fig.tight_layout()
    fig.savefig(out / "divergence_vs_error.png", dpi=150)
    plt.close(fig)


if __name__ == "__main__":
    main()
