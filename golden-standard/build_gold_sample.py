"""Build a blind, multi-label stratified sample for manual gold annotation.

The ambiguity study (PDF section 4) validates the three cloud LLMs against a
human gold standard built on a *stratified* sample that proportionally covers
all 16 categories. This script produces that sample.

Strategy
--------
The dataset is multi-label: every diff carries the set of categories scored
>= ``--threshold`` (0.5) by the primary extraction model (gemma). Because
labels co-occur (~2.9 per diff), strata overlap, so plain per-stratum
sampling does not work. Instead we:

  1. **Census the rarest labels** (``--census`` , default UI/UX + Performance):
     take *every* diff that carries one of them, since they are too rare to
     sample reliably.
  2. **Iterative stratification** (Sechidis et al. 2011) over the rest, to
     reach ``--size`` diffs while preserving each label's proportion across
     all 16 categories simultaneously.

gemma's predictions are used ONLY to stratify — they are NOT the gold, and never
appear on the sheet, so BOTH annotation axes stay blind. The annotator fills two
empty columns from scratch: ``labels_gold`` (the applicable ACF categories) and
``maintenance_gold`` (the ISO/IEC/IEEE 14764 maintenance TYPE —
corrective/preventive/adaptive/additive/perfective). Only the maintenance *type*
is a gold target; the parent class is derivable. Both axes live on the SAME
sheet, so a run produces the classic 4 files.

Re-running is safe: if a sample already exists in ``--out-dir`` the SAME diff_ids
are reused and any human-filled gold (both axes) is carried over, so a rebuild —
e.g. to add new columns — never discards annotation work. Pass ``--resample`` to
draw a fresh sample instead.

Outputs (under ``--out-dir``) — always these four
-------------------------------------------------
  * ``gold_sample_blind.csv``   - one row per diff. Columns: diff_id, filename,
    commit_url, commit_date, commit_message, prev_change_1/2/3 (the up-to-3
    earlier edits of this file, oldest first, each shown as ``date — message``
    and — in the .xlsx — hyperlinked to that commit: the SAME maintenance-reason
    trajectory the model is fed, and enough to see the file's prior state),
    changed_lines, ``labels_gold`` (empty; fill with ALL applicable categories)
    and ``maintenance_gold`` (empty; fill with the maintenance type). The
    prev_change_* links are what disambiguate adaptive-correction vs
    -enhancement and corrective vs perfective. In the .xlsx, commit_url and
    prev_change_* are clickable links.
  * ``gold_sample_blind.xlsx`` - Excel-friendly copy of the same sheet.
  * ``gold_sample_blind.jsonl`` - machine-readable twin (``labels_gold: []``).
  * ``gold_sample_key.jsonl``   - NOT for annotators: the gemma category label
    set / primary / census flag and the gemma maintenance prediction (leaf,
    base_type, class) per diff, plus the RNG seed, for scoring.

Example
-------
    python build_gold_sample.py --size 150 --seed 42
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

# This phase lives in its own folder but reuses the acf-analysis helpers and
# the extraction outputs, so we resolve those sibling paths explicitly.
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
ACF_ANALYSIS_DIR = REPO_ROOT / "acf-analysis"
sys.path.insert(0, str(ACF_ANALYSIS_DIR))
from acf_io import load_diffs_from_acf_json  # noqa: E402

# Canonical outputs location: acf-analysis/acf-outputs (same as run_multi_model.py
# and agreement_analysis.py). Previously this pointed at REPO_ROOT/acf-outputs,
# which split the primary (gemma) run from the study models under acf-analysis/.
DEFAULT_GEMMA_DIR = ACF_ANALYSIS_DIR / "acf-outputs" / "gemma4_31b-cloud"
DEFAULT_DIFFS = REPO_ROOT / "outputs" / "git_history_diffs.json"
DEFAULT_OUT_DIR = SCRIPT_DIR / "gold-sample"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument(
        "--gemma-file",
        type=Path,
        default=None,
        help="primary_*.jsonl from the extraction model (default: latest in "
             f"{DEFAULT_GEMMA_DIR}).",
    )
    p.add_argument("--diffs-json", type=Path, default=DEFAULT_DIFFS,
                   help=f"Source diffs (raw text), default {DEFAULT_DIFFS}.")
    p.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR,
                   help=f"Output directory, default {DEFAULT_OUT_DIR}.")
    p.add_argument("--threshold", type=float, default=0.5,
                   help="Score >= threshold marks a category as present (default 0.5).")
    p.add_argument("--size", type=int, default=150,
                   help="Target number of diffs in the sample (default 150).")
    p.add_argument("--census", nargs="*", default=["UI/UX", "Performance"],
                   help="Labels whose every diff is taken in full (default: UI/UX Performance).")
    p.add_argument("--seed", type=int, default=42, help="RNG seed for reproducibility.")
    p.add_argument(
        "--resample",
        action="store_true",
        help="Force a fresh sample even if a previous gold sample exists in "
             "--out-dir. By default, when an existing sample is found its diff_ids "
             "are reused (same diffs) and any human-filled gold labels are carried "
             "over, so re-running to add columns never discards annotation work.",
    )
    p.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip the automatic snapshot. By default, before overwriting, any "
             "existing gold_sample_* files are copied into a timestamped "
             "``_autobackup_<ts>`` folder inside --out-dir, so a mistaken run "
             "(e.g. --resample) is always recoverable.",
    )
    return p.parse_args()


def find_latest_primary(model_dir: Path) -> Path | None:
    if not model_dir.exists():
        return None
    cands = sorted(model_dir.glob("primary_*.jsonl"), key=lambda x: x.stem, reverse=True)
    return cands[0] if cands else None


def cleaned_changed_lines(diff_text: str) -> str:
    """Return the added/removed lines, stripped of diff markers, one per line."""
    out: list[str] = []
    for line in (diff_text or "").splitlines():
        if line.startswith("+++") or line.startswith("---"):
            continue
        if line.startswith("@@"):
            continue
        if line.startswith("+"):
            body = line[1:].strip()
            if body:
                out.append(f"+ {body}")
        elif line.startswith("-"):
            body = line[1:].strip()
            if body:
                out.append(f"- {body}")
    return "\n".join(out)


def label_set(scores: dict[str, float], threshold: float) -> list[str]:
    return sorted(c for c, v in scores.items() if isinstance(v, (int, float)) and v >= threshold)


def commit_file_url(repo: str, commit_hash: str, filename: str) -> str:
    """Link to a commit, anchored on the diff of *filename* only.

    GitHub has no standalone "single file inside a commit" page, but it anchors
    each file's diff as ``#diff-<sha256(path)>``, so the link opens the commit
    and jumps straight to the analysed file. If the anchor ever fails to match
    (e.g. GitHub changes the scheme) it degrades gracefully to the commit top.
    """
    if not (repo and commit_hash):
        return ""
    url = f"https://github.com/{repo}/commit/{commit_hash}"
    if filename:
        anchor = hashlib.sha256(filename.encode("utf-8")).hexdigest()
        url += f"#diff-{anchor}"
    return url


# Match acf_analyser: the model is fed at most this many earlier edits to the
# same file as the maintenance-reason trajectory. Keep in sync with _MAX_PRIOR.
_MAX_PRIOR_CHANGES = 3


def build_prior_changes(diffs) -> dict[str, list[dict[str, str]]]:
    """Map each diff_id to the up-to-3 EARLIER edits of the same (repo, filename),
    oldest first — the exact maintenance-reason trajectory the model receives
    (see acf_analyser). Each entry is ``{commit_date, commit_message,
    commit_hash, commit_url}`` (the url links the earlier commit's file diff).

    commit_date is ISO-8601 so a string sort equals chronological order; diffs
    with a missing date sort first, mirroring the analyser.
    """
    by_file: dict[tuple[str, str], list[dict]] = {}
    for d in diffs:
        key = (str(d.get("repo", "")), str(d.get("filename", "unknown")))
        by_file.setdefault(key, []).append(d)
    out: dict[str, list[dict[str, str]]] = {}
    for items in by_file.values():
        ordered = sorted(items, key=lambda x: str(x.get("timestamp", "")))
        history: list[dict[str, str]] = []
        for d in ordered:
            out[str(d.get("diff_id", ""))] = history[-_MAX_PRIOR_CHANGES:]
            repo = str(d.get("repo", "") or "").strip()
            chash = str(d.get("commit_hash", "") or "").strip()
            fname = str(d.get("filename", "") or "").strip()
            history = history + [{
                "commit_date": str(d.get("timestamp", "")),
                "commit_message": str(d.get("commit_message", "")),
                "commit_hash": chash,
                "commit_url": commit_file_url(repo, chash, fname),
            }]
    return out


def prior_change_display(entry: dict[str, str], max_msg: int = 80) -> str:
    """Cell text for one earlier edit: ``YYYY-MM-DD — message`` (message trimmed).

    Shown as the visible label of a hyperlink to that commit, so the annotator
    reads the trajectory at a glance and can click through for the full diff.
    """
    date = (entry.get("commit_date") or "").strip()
    day = date[:10] if len(date) >= 10 else date
    msg = " ".join((entry.get("commit_message") or "").split())
    if len(msg) > max_msg:
        msg = msg[: max_msg - 1].rstrip() + "…"
    return f"{day} — {msg}".strip(" —")


def split_labels(value: str) -> list[str]:
    """Parse a human-typed gold cell (``"Testing; Documentation"``) into a list.

    Tolerates ';' or ',' separators and stray whitespace; order is preserved and
    duplicates dropped so it round-trips cleanly with ``join_labels``.
    """
    if not isinstance(value, str):
        return []
    parts = [p.strip() for p in re.split(r"[;,]", value) if p.strip()]
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def join_labels(labels) -> str:
    """Inverse of ``split_labels`` for the CSV/XLSX cell (``"A; B"``)."""
    if isinstance(labels, str):
        return labels.strip()
    if isinstance(labels, (list, tuple)):
        return "; ".join(str(x).strip() for x in labels if str(x).strip())
    return ""


def _rows_from_xlsx(path: Path) -> tuple[list[str], list[list[str]]] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    if not path.exists():
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [str(c) if c is not None else "" for c in rows[0]]
    body = [[("" if c is None else str(c)) for c in r] for r in rows[1:]]
    return header, body


def _rows_from_csv(path: Path) -> tuple[list[str], list[list[str]]] | None:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def load_existing_annotations(
    out_dir: Path, stem: str, value_cols: tuple[str, ...], id_col: str = "diff_id"
) -> dict[str, dict[str, str]]:
    """Recover human-filled gold values for a sheet, keyed by diff_id.

    Reads ``{stem}.xlsx`` then ``{stem}.csv`` then ``{stem}.jsonl`` (in that
    order of trust — Excel users usually edit the .xlsx) and MERGES them: a
    diff_id's value for a column is filled by the first source that has a
    non-empty entry, so it does not matter which file the annotator edited.
    Returns ``{diff_id: {col: value}}``; only non-empty values are recorded.
    """
    result: dict[str, dict[str, str]] = {}

    def absorb(header: list[str], body: list[list[str]]) -> None:
        if id_col not in header:
            return
        idx_id = header.index(id_col)
        col_idx = {c: header.index(c) for c in value_cols if c in header}
        for row in body:
            if idx_id >= len(row):
                continue
            did = (row[idx_id] or "").strip()
            if not did:
                continue
            bucket = result.setdefault(did, {})
            for col, ci in col_idx.items():
                if col in bucket:  # first non-empty source wins
                    continue
                val = row[ci].strip() if ci < len(row) and isinstance(row[ci], str) else ""
                if val:
                    bucket[col] = val

    for reader, name in ((_rows_from_xlsx, f"{stem}.xlsx"), (_rows_from_csv, f"{stem}.csv")):
        parsed = reader(out_dir / name)
        if parsed is not None:
            absorb(*parsed)

    # JSONL fallback: labels_gold may be a list there, so stringify to match.
    jpath = out_dir / f"{stem}.jsonl"
    if jpath.exists():
        for line in jpath.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            did = str(rec.get(id_col, "")).strip()
            if not did:
                continue
            bucket = result.setdefault(did, {})
            for col in value_cols:
                if col in bucket or col not in rec:
                    continue
                raw = rec[col]
                val = join_labels(raw) if isinstance(raw, (list, tuple)) else str(raw or "").strip()
                if val:
                    bucket[col] = val

    return result


def load_existing_sample_ids(out_dir: Path) -> list[str]:
    """Return the diff_ids of a previously built blind sample, in file order.

    Prefers the JSONL (authoritative, machine-written); falls back to the CSV.
    Empty list when no prior sample exists.
    """
    for reader, name in (
        (None, "gold_sample_blind.jsonl"),
        (_rows_from_csv, "gold_sample_blind.csv"),
    ):
        path = out_dir / name
        if not path.exists():
            continue
        if name.endswith(".jsonl"):
            ids: list[str] = []
            seen: set[str] = set()
            for line in path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    did = str(json.loads(line).get("diff_id", "")).strip()
                except json.JSONDecodeError:
                    continue
                if did and did not in seen:
                    seen.add(did)
                    ids.append(did)
            if ids:
                return ids
        else:
            parsed = reader(path)
            if parsed and "diff_id" in parsed[0]:
                idx = parsed[0].index("diff_id")
                ids, seen = [], set()
                for row in parsed[1]:
                    did = (row[idx].strip() if idx < len(row) else "")
                    if did and did not in seen:
                        seen.add(did)
                        ids.append(did)
                if ids:
                    return ids
    return []


def backup_existing_outputs(out_dir: Path) -> Path | None:
    """Snapshot any existing ``gold_sample_*`` files before a rebuild overwrites
    them, so a mistaken run is always recoverable.

    Copies them into ``out_dir/_autobackup_<timestamp>/``. That prefix keeps the
    snapshot out of the way of the reuse/carry-over readers (which look only at
    the top-level ``gold_sample_*`` files, never inside subfolders). Returns the
    backup directory, or ``None`` when there was nothing to save.
    """
    existing = sorted(p for p in out_dir.glob("gold_sample_*") if p.is_file())
    if not existing:
        return None
    dest = out_dir / f"_autobackup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    dest.mkdir(parents=True, exist_ok=True)
    for p in existing:
        shutil.copy2(p, dest / p.name)
    return dest


def write_xlsx_from_csv(
    csv_path: Path,
    xlsx_path: Path,
    widths: dict[int, int] | None = None,
    wrap_cols: tuple[int, ...] = (4, 5),
    title: str = "gold_sample",
    hyperlinks: dict[tuple[int, int], str] | None = None,
) -> bool:
    """Render a blind CSV as a real .xlsx so it opens cleanly in Excel.

    Solves the two Excel pain points of a raw CSV: locale-dependent delimiter
    (Italian Excel splits on ';', not ',') and multi-line cells. ``widths`` maps
    1-based column index -> width, ``wrap_cols`` are the columns whose long text
    should wrap. ``hyperlinks`` maps a 1-based ``(row, col)`` (row 1 is the
    header) to a URL; that cell keeps its CSV text as the visible label but
    becomes a clickable link — used for the commit links and per-earlier-edit
    ``prev_change_*`` cells. Returns False if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("[xlsx] openpyxl not installed; skipping .xlsx "
              "(install with: pip install openpyxl)", file=sys.stderr)
        return False

    # Diff text can carry control characters that are illegal in the XLSX XML;
    # Excel flags the file as corrupt and "repairs" it. Strip them up front.
    def clean(value: str) -> str:
        return ILLEGAL_CHARACTERS_RE.sub("", value) if isinstance(value, str) else value

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    wrap_set = set(wrap_cols)
    links = hyperlinks or {}

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=clean(value))
            if r_idx == 1:
                cell.font = header_font
            else:
                cell.alignment = wrap if c_idx in wrap_set else top
                url = links.get((r_idx, c_idx))
                if url and value:
                    cell.hyperlink = url
                    cell.font = link_font

    # Default widths match the base blind sheet:
    # diff_id, filename, commit_url, commit_message, changed_lines, labels_gold
    if widths is None:
        widths = {1: 42, 2: 20, 3: 60, 4: 40, 5: 90, 6: 26}
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    wb.save(xlsx_path)
    return True


def iterative_stratification(
    pool: list[str],
    labels: dict[str, list[str]],
    size: int,
    rng: random.Random,
) -> list[str]:
    """Single-subset iterative stratification (Sechidis et al. 2011).

    Greedily pick ``size`` ids from ``pool`` so each label's proportion is
    preserved as closely as possible. At each step we serve the label that is
    most "starved" (smallest remaining desired count), breaking ties toward the
    rarest label, then take the diff that best fills the still-wanted labels.
    """
    pool_set = set(pool)
    if size <= 0 or not pool_set:
        return []
    if size >= len(pool_set):
        return list(pool_set)

    ratio = size / len(pool_set)
    label_total: Counter[str] = Counter()
    for i in pool_set:
        for lab in labels[i]:
            label_total[lab] += 1
    desired: dict[str, float] = {lab: ratio * tot for lab, tot in label_total.items()}

    selected: list[str] = []
    while len(selected) < size and pool_set:
        wanted = [lab for lab in desired if desired[lab] > 0 and label_total[lab] > 0]
        if not wanted:
            # No label still needs samples: fill the remainder uniformly.
            rest = list(pool_set)
            rng.shuffle(rest)
            for i in rest[: size - len(selected)]:
                selected.append(i)
                pool_set.discard(i)
            break

        min_des = min(desired[lab] for lab in wanted)
        tied = [lab for lab in wanted if desired[lab] == min_des]
        min_tot = min(label_total[lab] for lab in tied)
        tied = [lab for lab in tied if label_total[lab] == min_tot]
        lab = rng.choice(sorted(tied))

        cands = [i for i in pool_set if lab in labels[i]]
        rng.shuffle(cands)  # randomise ties under the max() below
        best = max(cands, key=lambda i: sum(max(0.0, desired[l]) for l in labels[i]))

        selected.append(best)
        pool_set.discard(best)
        for l in labels[best]:
            desired[l] -= 1
            label_total[l] -= 1

    return selected


def load_gemma_maintenance(gemma_file: Path) -> dict[str, dict[str, str]]:
    """Return ``{diff_id: {leaf, base_type, class, rationale}}`` from the primary
    extraction file's ``maintenance`` block (the LLM's modification-request call)."""
    out: dict[str, dict[str, str]] = {}
    for line in gemma_file.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        did = rec.get("diff_id")
        if not did:
            continue
        m = rec.get("maintenance") or {}
        out[did] = {
            "leaf": str(m.get("primary_maintenance_type") or ""),
            "base_type": str(m.get("maintenance_base_type") or ""),
            "class": str(m.get("maintenance_class") or ""),
        }
    return out


def main() -> None:
    args = parse_args()
    rng = random.Random(args.seed)

    gemma_file = args.gemma_file or find_latest_primary(DEFAULT_GEMMA_DIR)
    if not gemma_file or not gemma_file.exists():
        sys.exit(f"Extraction file not found (looked in {DEFAULT_GEMMA_DIR}).")
    if not args.diffs_json.exists():
        sys.exit(f"Source diffs not found: {args.diffs_json}")

    # 1. gemma label sets (stratification variable only).
    labels: dict[str, list[str]] = {}
    gemma_primary: dict[str, str] = {}
    for line in gemma_file.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        did = rec.get("diff_id")
        if not did:
            continue
        labels[did] = label_set(rec.get("category_scores", {}), args.threshold)
        gemma_primary[did] = (rec.get("primary") or {}).get("primary_category", "")

    # 2. raw diff text + commit message, joined by diff_id.
    src = {d["diff_id"]: d for d in load_diffs_from_acf_json(args.diffs_json)}

    # Drop diffs with no label above threshold or no source text: nothing to stratify on.
    usable = [d for d in labels if labels[d] and d in src]
    dropped = len(labels) - len(usable)
    if dropped:
        print(f"[info] skipped {dropped} diff(s) with empty label set or missing source text")

    # 3. Selection: reuse the previous sample (default, so annotation work is
    #    preserved) or draw a fresh stratified one when --resample is given.
    census_set = set(args.census)
    reused_ids = [] if args.resample else load_existing_sample_ids(args.out_dir)
    reused_ids = [d for d in reused_ids if d in src]  # keep only renderable diffs

    if reused_ids:
        selected = reused_ids
        censused_ids = {d for d in selected if census_set.intersection(labels.get(d, []))}
        censused = [d for d in selected if d in censused_ids]
        strat = [d for d in selected if d not in censused_ids]
        print(f"[reuse] existing sample found -> reusing {len(selected)} diff_ids "
              f"(same diffs; pass --resample to draw a new sample)")
        absent = [d for d in selected if d not in labels]
        if absent:
            print(f"[reuse] {len(absent)} reused diff(s) are absent from the current "
                  f"extraction file; kept anyway with blank gemma labels.")
    else:
        # 3a. Census the rarest labels.
        censused = [d for d in usable if census_set.intersection(labels[d])]
        censused_ids = set(censused)
        if len(censused) > args.size:
            sys.exit(
                f"Census of {sorted(census_set)} already yields {len(censused)} diffs "
                f"> --size {args.size}. Raise --size or narrow --census."
            )

        # 3b. Iterative stratification on the rest.
        remaining_pool = [d for d in usable if d not in censused_ids]
        need = args.size - len(censused)
        strat = iterative_stratification(remaining_pool, labels, need, rng)

        selected = censused + strat
        rng.shuffle(selected)  # mix census + stratified so the sheet order is not informative

    # 3c. Maintenance-reason prediction (the "why" axis) the primary model now
    #     also emits — verified in the SAME sheet, so a run stays 4 files.
    maint = load_gemma_maintenance(gemma_file)

    # 3c-bis. Per-file change trajectory: the SAME earlier-edits context the model
    #     receives for the maintenance axis (acf_analyser -> prompt rule 16), so a
    #     human judges the reason on equal footing. Computed over ALL diffs.
    prior_by_id = build_prior_changes(src.values())

    # 3d. Carry over any human-filled gold from a previous build so a re-run
    #     (e.g. to add new columns) never discards annotation work. Covers both
    #     the category labels and the maintenance-type verification.
    carried = load_existing_annotations(
        args.out_dir, "gold_sample_blind",
        ("labels_gold", "maintenance_gold"),
    )
    # Migration bridge: recover the maintenance gold from the legacy, standalone
    # ``gold_sample_maintenance.*`` sheet (the pre-merge 8-file layout) so
    # switching to the merged sheet does not lose that work. The merged sheet
    # always wins when both carry a value.
    for did, vals in load_existing_annotations(
        args.out_dir, "gold_sample_maintenance",
        ("maintenance_gold",),
    ).items():
        bucket = carried.setdefault(did, {})
        for col, val in vals.items():
            bucket.setdefault(col, val)

    carried_labels = {
        did: v["labels_gold"] for did, v in carried.items() if v.get("labels_gold")
    }

    # 4. Coverage report.
    cov = Counter()
    for d in selected:
        for lab in labels.get(d, []):
            cov[lab] += 1
    all_labels = sorted({lab for ls in labels.values() for lab in ls})
    print(f"\n[sample] gemma file : {gemma_file.name}")
    print(f"[sample] threshold  : {args.threshold}")
    print(f"[sample] size       : {len(selected)} "
          f"(census={len(censused)} from {sorted(census_set)}, stratified={len(strat)})")
    print(f"[sample] seed       : {args.seed}")
    if carried_labels:
        n_here = sum(1 for d in selected if d in carried_labels)
        print(f"[sample] carried over {n_here} human-filled labels_gold from a prior build")
    print("\n[coverage] label-instances in the sample (>= threshold):")
    for lab in sorted(all_labels, key=lambda x: cov[x], reverse=True):
        print(f"  {lab:22s} {cov[lab]:4d}")

    # 5. Write outputs (snapshot the previous ones first so a mistaken run is
    #    always recoverable — carry-over above has already read them).
    args.out_dir.mkdir(parents=True, exist_ok=True)
    if not args.no_backup:
        bdir = backup_existing_outputs(args.out_dir)
        if bdir is not None:
            print(f"[backup] snapshot of previous outputs -> {bdir}")
    blind_csv = args.out_dir / "gold_sample_blind.csv"
    blind_jsonl = args.out_dir / "gold_sample_blind.jsonl"
    key_jsonl = args.out_dir / "gold_sample_key.jsonl"

    # Single sheet, two blind axes: category labels + maintenance type. The
    # maintenance-context columns give the human the SAME basis the model used to
    # judge the reason. commit_url and prev_change_{1,2,3} render as clickable
    # links in the .xlsx; prev_change_* are the up-to-3 earlier edits of this
    # file (oldest first), each linking that commit's file diff — these already
    # cover the file's prior state, so no separate parent-commit link is needed.
    #   1 diff_id  2 filename  3 commit_url  4 commit_date  5 commit_message
    #   6 prev_change_1  7 prev_change_2  8 prev_change_3  9 changed_lines
    #   10 labels_gold  11 maintenance_gold
    header = [
        "diff_id", "filename", "commit_url", "commit_date", "commit_message",
        "prev_change_1", "prev_change_2", "prev_change_3",
        "changed_lines", "labels_gold", "maintenance_gold",
    ]
    # 1-based column indices that carry a clickable URL in the .xlsx.
    COL_COMMIT_URL = 3
    COL_PREV_CHANGE = (6, 7, 8)  # prev_change_1..3
    hyperlinks: dict[tuple[int, int], str] = {}

    n_maint_missing = 0
    with (
        blind_csv.open("w", encoding="utf-8-sig", newline="") as fcsv,
        blind_jsonl.open("w", encoding="utf-8") as fjsonl,
        key_jsonl.open("w", encoding="utf-8") as fkey,
    ):
        writer = csv.writer(fcsv)
        writer.writerow(header)
        for r_idx, did in enumerate(selected, start=2):  # row 1 is the header
            d = src[did]
            changed = cleaned_changed_lines(d.get("diff_text", ""))
            filename = d.get("filename", "")
            commit_message = (d.get("commit_message", "") or "").strip()
            commit_date = (d.get("timestamp", "") or "").strip()
            repo = (d.get("repo", "") or "").strip()
            commit_hash = (d.get("commit_hash", "") or "").strip()
            commit_url = commit_file_url(repo, commit_hash, filename)

            # Up to 3 earlier edits (oldest first) -> one cell each, showing
            # "date — message" and linking that commit.
            prior_items = prior_by_id.get(did, [])
            prev_cells = ["", "", ""]
            for i, entry in enumerate(prior_items[:3]):
                prev_cells[i] = prior_change_display(entry)
                url = entry.get("commit_url", "")
                if url:
                    hyperlinks[(r_idx, COL_PREV_CHANGE[i])] = url

            if commit_url:
                hyperlinks[(r_idx, COL_COMMIT_URL)] = commit_url

            # The model's maintenance prediction is kept in the key only (blind
            # sheet), but we still track how many diffs the model covered.
            m = maint.get(did)
            if m is None:
                n_maint_missing += 1
                m = {"leaf": "", "base_type": "", "class": ""}

            ann = carried.get(did, {})
            gold_cell = ann.get("labels_gold", "")     # category labels (blank if new)
            m_gold = ann.get("maintenance_gold", "")   # maintenance type (blank if new)

            writer.writerow([
                did, filename, commit_url, commit_date, commit_message,
                prev_cells[0], prev_cells[1], prev_cells[2],
                changed, gold_cell, m_gold,
            ])
            fjsonl.write(json.dumps({
                "diff_id": did,
                "filename": filename,
                "commit_url": commit_url,
                "commit_date": commit_date,
                "commit_message": commit_message,
                # Machine-readable trajectory (same shape the model receives,
                # plus commit_hash/commit_url for linking).
                "prior_changes": prior_items,
                "changed_lines": changed,
                "labels_gold": split_labels(gold_cell),
                "maintenance_gold": m_gold,
            }, ensure_ascii=False) + "\n")
            fkey.write(json.dumps({
                "diff_id": did,
                "gemma_labels": labels.get(did, []),
                "gemma_primary": gemma_primary.get(did, ""),
                "in_census": did in censused_ids,
                "seed": args.seed,
                # Maintenance prediction kept with the key (not for annotators),
                # incl. the derivable parent class for later scoring.
                "gemma_maintenance_leaf": m["leaf"],
                "gemma_maintenance_base_type": m["base_type"],
                "gemma_maintenance_class": m["class"],
            }, ensure_ascii=False) + "\n")

    # Excel-friendly copy derived from the CSV we just wrote. Wrap the long text
    # columns: commit_message=5, prev_change_1..3=6-8, changed_lines=9. Commit
    # link and prev_change_* cells are made clickable via the hyperlinks map.
    blind_xlsx = args.out_dir / "gold_sample_blind.xlsx"
    xlsx_widths = {
        1: 42, 2: 20, 3: 55, 4: 20, 5: 40,
        6: 38, 7: 38, 8: 38, 9: 90, 10: 26, 11: 18,
    }
    xlsx_ok = write_xlsx_from_csv(
        blind_csv, blind_xlsx, widths=xlsx_widths, wrap_cols=(5, 6, 7, 8, 9),
        hyperlinks=hyperlinks,
    )

    covered = len(selected) - n_maint_missing
    print(f"\n[maintenance] LLM type found for {covered}/{len(selected)} diffs"
          + (f"  ({n_maint_missing} missing)" if n_maint_missing else "")
          + " (kept in the key; the sheet stays blind)")
    print(f"\n[+] gold sheet (CSV)  : {blind_csv}")
    if xlsx_ok:
        print(f"[+] gold sheet (XLSX) : {blind_xlsx}")
    print(f"[+] gold sheet (JSONL): {blind_jsonl}")
    print(f"[+] key (NOT for annotators): {key_jsonl}")
    print(
        "\nCategory axis: fill 'labels_gold' with all applicable categories "
        "(primary included), e.g. \"Testing; Documentation\".\n"
        "Maintenance axis: fill 'maintenance_gold' with the correct type "
        "(corrective/preventive/adaptive/additive/perfective); click "
        "'prev_change_1..3' (earlier edits of this file — the same context the "
        "model saw, and the file's prior state) to disambiguate."
    )


if __name__ == "__main__":
    main()
